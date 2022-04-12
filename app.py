import openpyxl as xl
#  importing the barchart
from openpyxl.chart import BarChart, Reference

# loading the sheet
wb =xl.load_workbook('transactions.xlsx')

sheet = wb["Sheet1"]

# getting the cell of the sheet
cell = sheet["a1"]
# second way of getting the cell
# (row, column)
cell = sheet.cell(1,1)
# this is showing the value of the cell in row 1 column 1
print(cell.value)

# for loop for the rows
for row in range(2, sheet.max_row + 1):
# getting the value of the prices
    cell = sheet.cell(row,3)
    # correcting the price
    corrected_price = cell.value * 0.9
    # getting the actual refference to the cell
    corrected_price_cell = sheet.cell(row, 4)
    # changing the values of the cells
    corrected_price_cell.value = corrected_price
    print(corrected_price)


# this is the code to add the chart
values = Reference(sheet, min_row =2, max_row= sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')


